import argparse
import math
import os
import logging
from decimal import Decimal, ROUND_HALF_UP
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Color, Font

from ValueProcessor import ValuesHolder, ValuesHelper

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[
        logging.FileHandler("excel_processing.log", mode='w'),
        logging.StreamHandler()
    ]
)

ACCURACY = 0.01
MODEL_COL = 'B'
PRICE_EUR_COL = 'C'
PRICE_HRN_COL = 'D'
PERCENT_HRN_COL = 'E'
WEIGHT_PRICE_HRN_COL = 'F'
ADJUST_SUM_HRN_COL = 'G'
ADJUST_SUM_EUR_COL = 'H'
CHECK_EUR_COL = 'I'
START_COLUMN = 2
END_COLUMN = 8
START_DATA_ROW = 6
DATE_CELL = f"{MODEL_COL}{START_COLUMN + 1}"
EURO_RATE_CELL = f"{PRICE_EUR_COL}{START_COLUMN + 1}"
ADJUST_SUM_EUR_CELL = f"{ADJUST_SUM_EUR_COL}{START_COLUMN + 1}"
ADJUST_SUM_HRN_CELL = f"{ADJUST_SUM_HRN_COL}{START_COLUMN + 1}"

PALE_BLUE_FILL = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
PALE_GREEN_FILL = PatternFill(start_color="98FB98", end_color="98FB98", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
RED_FONT = Color('FF0000')
ORANGE_FONT = Color('EE9A00')
BLUE_FONT = Color('0000FF')

def parse_args():
    parser = argparse.ArgumentParser(description="Picker")
    parser.add_argument('--manual', action=argparse.BooleanOptionalAction)
    return parser.parse_args()

def _round(value, number=2):
    # value = round(round(round(value, 4), 3), 2)
    # value = round(value, 2)
    for i in range(4, number-1, -1):
        value = float(Decimal(str(value)).quantize(Decimal('1.' + '1' * i), rounding=ROUND_HALF_UP))
    return value

def mark_cell(sheet, col: str, row: int, color: Color):
    sheet[f'{col}{row}'].font = Font(color=color)


def write_cell(sheet, col: str, row: str, value):
    sheet[f'{col}{row}'].value = value


def read_cell(sheet, col: str, row: str):
    return sheet[f'{col}{row}'].value

def process_excel_file(file_path, manual):
    def count_non_empty_models():
        count = 0
        for row in range(START_DATA_ROW, ws.max_row + 1):
            if ws[f'{MODEL_COL}{row}'].value is not None:  # Check if the cell is not empty
                count += 1
            else:
                break  # Stop counting at the first empty cell
        return count

    def is_sums_equal():
        return _round(sum_hrn) == _round(target_sum_hrn) and _round(sum_eur) == _round(target_sum_eur)

    def calc_sums():
        s_hrn = sum(_round(ws[f'{ADJUST_SUM_HRN_COL}{row}'].value) for row in range(START_DATA_ROW, START_DATA_ROW + non_empty_rows))
        s_eur = sum(_round(ws[f'{ADJUST_SUM_EUR_COL}{row}'].value) for row in range(START_DATA_ROW, START_DATA_ROW + non_empty_rows))
        return s_hrn, s_eur

    # Load the Excel file with data_only=True to get cell values instead of formulas
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    try:
        # date = ws[date_cell].value
        euro_rate = ws[EURO_RATE_CELL].value
        target_sum_hrn = ws[ADJUST_SUM_HRN_CELL].value
        non_empty_rows = count_non_empty_models()
        target_sum_eur = target_sum_hrn / euro_rate
        sum_price_models_euro = sum(ws[f'{PRICE_EUR_COL}{row}'].value for row in range(START_DATA_ROW, START_DATA_ROW + non_empty_rows))
        recalculated_diff_euro = target_sum_eur - sum_price_models_euro
        # write calculated Target EUR value into the sheet
        ws[ADJUST_SUM_EUR_CELL].value = target_sum_eur if manual else _round(target_sum_eur)
    except TypeError:
        logging.error(f"Error reading values from the worksheet '{file_path}'. Check that the data is correct.")
        return

    if sum(ws[f'{PERCENT_HRN_COL}{row}'].value for row in range(START_DATA_ROW, START_DATA_ROW + non_empty_rows)) != 100:
        logging.error(f"Total weight percent in column '{PERCENT_HRN_COL}' of file {file_path} does not equal 100%.")
        return

    # main calculations
    for row in range(START_DATA_ROW, START_DATA_ROW + non_empty_rows):
        price_euro = ws[f'{PRICE_EUR_COL}{row}'].value
        price_model_hrn = price_euro * euro_rate
        weight_percent = ws[f'{PERCENT_HRN_COL}{row}'].value
        sum_by_percent = recalculated_diff_euro * euro_rate * weight_percent / 100

        ws[f'{PRICE_HRN_COL}{row}'] = price_model_hrn
        ws[f'{WEIGHT_PRICE_HRN_COL}{row}'] = abs(sum_by_percent)
        ws[f'{ADJUST_SUM_HRN_COL}{row}'] = ws[f'{PRICE_HRN_COL}{row}'].value + sum_by_percent
        ws[f'{ADJUST_SUM_EUR_COL}{row}'] = ws[f'{ADJUST_SUM_HRN_COL}{row}'].value / euro_rate
        ws[f'{ADJUST_SUM_HRN_COL}{row + non_empty_rows + 2}'] = ws[f'{PRICE_HRN_COL}{row}'].value + sum_by_percent

    if not manual:
        adjustment(ws, non_empty_rows, euro_rate)

    sum_hrn, sum_eur = calc_sums()

    if not is_sums_equal():
        logging.error(
            f"Could not reach required accuracy for file: {file_path}. "
            f"Expected sum for HRN: {_round(target_sum_hrn)}, Calculated: {_round(sum_hrn)}. "
            f"Expected sum for EUR: {_round(target_sum_eur)}, Calculated: {_round(sum_eur)}."
        )

    summary_row = START_DATA_ROW + non_empty_rows
    ws[f'{ADJUST_SUM_HRN_COL}{summary_row}'] = sum_hrn if manual else _round(sum_hrn)
    ws[f'{ADJUST_SUM_EUR_COL}{summary_row}'] = sum_eur if manual else _round(sum_eur)

    # rounding HRN & EUR columns
    for row in range(START_DATA_ROW, START_DATA_ROW + non_empty_rows):
        if not manual:
            ws[f'{PRICE_HRN_COL}{row}'].value = _round(ws[f'{PRICE_HRN_COL}{row}'].value)
            ws[f'{WEIGHT_PRICE_HRN_COL}{row}'].value = _round(ws[f'{WEIGHT_PRICE_HRN_COL}{row}'].value)
            ws[f'{ADJUST_SUM_HRN_COL}{row}'].value = _round(ws[f'{ADJUST_SUM_HRN_COL}{row}'].value)
            ws[f'{ADJUST_SUM_EUR_COL}{row}'].value = _round(ws[f'{ADJUST_SUM_EUR_COL}{row}'].value)
        ws[f'{ADJUST_SUM_EUR_COL}{row + non_empty_rows + 2}'].value = ws[f'{ADJUST_SUM_HRN_COL}{row}'].value
        ws[f'{CHECK_EUR_COL}{row + non_empty_rows + 2}'] = f'={ADJUST_SUM_EUR_COL}{row + non_empty_rows + 2}/{EURO_RATE_CELL}'
        ws[f'{CHECK_EUR_COL}{row + non_empty_rows + 2}'].number_format = '0.00000'
        mark_cell(ws, CHECK_EUR_COL, row + non_empty_rows + 2, BLUE_FONT)

    row_sum = START_DATA_ROW + non_empty_rows * 2 + 3
    row_start = START_DATA_ROW + non_empty_rows + 2
    row_end = START_DATA_ROW + non_empty_rows + non_empty_rows + 1
    ws[f'{ADJUST_SUM_HRN_COL}{row_sum}'].value = f'=SUM({ADJUST_SUM_HRN_COL}{row_start}:{ADJUST_SUM_HRN_COL}{row_end})'
    ws[f'{ADJUST_SUM_EUR_COL}{row_sum}'].value = f'=SUM({ADJUST_SUM_EUR_COL}{row_start}:{ADJUST_SUM_EUR_COL}{row_end})'
    ws[f'{CHECK_EUR_COL}{row_sum}'].value = f'=SUM({CHECK_EUR_COL}{row_start}:{CHECK_EUR_COL}{row_end})'
    mark_cell(ws, ADJUST_SUM_HRN_COL, row_sum, BLUE_FONT)
    mark_cell(ws, ADJUST_SUM_EUR_COL, row_sum, BLUE_FONT)
    mark_cell(ws, CHECK_EUR_COL, row_sum, BLUE_FONT)

    styling(ws, summary_row, non_empty_rows)

    try:
        wb.save(file_path)
    except PermissionError:
        logging.error(f"Unable to save the workbook: {file_path}. Check to see if it's open.")
        return

    logging.info(f"Processing of file {file_path} completed successfully.")


def adjustment(ws, non_empty_rows, euro_rate):
    def fill_values_holder_list(lst_hrn, ascending=True):
        def sum_of_fractions(value):
            fraction_str = f"{value:.6f}".split(".")[1]
            return sum(map(int, fraction_str))

        fractions = [math.modf(value)[0] for value in lst_hrn]
        sorted_indices = sorted(
            range(len(fractions)),
            key=lambda i: sum_of_fractions(fractions[i]),
            reverse=not ascending
        )
        holder = [ValuesHolder(sorted_indices.index(i), i, lst_hrn[i]) for i in range(len(lst_hrn))]
        return holder

    def calculate_values(col, cell):
        values = [_round(ws[f'{col}{r}'].value) for r in range(START_DATA_ROW, START_DATA_ROW + non_empty_rows)]
        target_sum = ws[f'{cell}'].value
        current_sum = sum(values)
        diff = _round(target_sum - current_sum)
        return values, target_sum, current_sum, diff

    values_hrn, target_sum_hrn, current_sum_hrn, diff_hrn = calculate_values(ADJUST_SUM_HRN_COL, ADJUST_SUM_HRN_CELL)
    values_eur, target_sum_eur, current_sum_eur, diff_eur = calculate_values(ADJUST_SUM_EUR_COL, ADJUST_SUM_EUR_CELL)

    values_helper = ValuesHelper(fill_values_holder_list(values_hrn, diff_hrn > 0))
    increment = ACCURACY if diff_eur > 0 else -ACCURACY

    while diff_eur != 0 or diff_hrn != 0:
        if diff_eur != 0:
            current_value_hrn = values_helper.next_min() if diff_eur > 0 else values_helper.next_max()
            current_value_eur = _round(read_cell(ws, ADJUST_SUM_EUR_COL, START_DATA_ROW + values_helper.index()))
            current_row = START_DATA_ROW + values_helper.index()
            new_value_eur = current_value_eur
            old_value_eur = current_value_eur

            while diff_eur != 0:
                current_value_hrn += increment
                new_value_eur = _round(current_value_hrn / euro_rate)
                current_sum_eur = _round(current_sum_eur - current_value_eur + new_value_eur)
                current_value_eur = new_value_eur
                diff_eur = _round(target_sum_eur - current_sum_eur)

            write_cell(ws, ADJUST_SUM_HRN_COL, current_row, current_value_hrn)
            write_cell(ws, ADJUST_SUM_EUR_COL, current_row, current_value_eur)
            mark_cell(ws, ADJUST_SUM_HRN_COL, current_row, RED_FONT)
            if new_value_eur != old_value_eur:
                mark_cell(ws, ADJUST_SUM_EUR_COL, current_row, RED_FONT)

        values_hrn, target_sum_hrn, current_sum_hrn, diff_hrn = calculate_values(ADJUST_SUM_HRN_COL, ADJUST_SUM_HRN_CELL)
        if diff_hrn != 0:
            current_value_hrn = values_helper.next_max() if diff_hrn > 0 else values_helper.next_min()
            current_row = START_DATA_ROW + values_helper.index()
            write_cell(ws, ADJUST_SUM_HRN_COL, current_row, current_value_hrn + diff_hrn)
            current_value_eur = _round(read_cell(ws, ADJUST_SUM_EUR_COL, current_row))
            new_value_eur = _round(current_value_hrn / euro_rate)
            old_value_eur = current_value_eur
            if current_value_eur != new_value_eur:
                logging.warning(f"Value EUR is changed due to value HRN adjustment. Was: {current_value_eur}, now: {new_value_eur}")
            mark_cell(ws, ADJUST_SUM_HRN_COL, current_row, ORANGE_FONT)
            if current_value_eur != old_value_eur:
                mark_cell(ws, ADJUST_SUM_EUR_COL, current_row, ORANGE_FONT)

        values_hrn, target_sum_hrn, current_sum_hrn, diff_hrn = calculate_values(ADJUST_SUM_HRN_COL, ADJUST_SUM_HRN_CELL)
        values_eur, target_sum_eur, current_sum_eur, diff_eur = calculate_values(ADJUST_SUM_EUR_COL, ADJUST_SUM_EUR_CELL)

def styling(ws, summary_row, non_empty_rows):
    ws[ADJUST_SUM_HRN_CELL].fill = PALE_GREEN_FILL
    ws[ADJUST_SUM_EUR_CELL].fill = PALE_BLUE_FILL
    ws[f'{ADJUST_SUM_HRN_COL}{summary_row}'].fill = PALE_GREEN_FILL
    ws[f'{ADJUST_SUM_EUR_COL}{summary_row}'].fill = PALE_BLUE_FILL

    for row in range(START_DATA_ROW, START_DATA_ROW + non_empty_rows):
        ws[f'{ADJUST_SUM_HRN_COL}{row}'].fill = YELLOW_FILL

    thick_border_left = Border(left=Side(style="thick"))
    thick_border_right = Border(right=Side(style="thick"))

    for row in range(START_DATA_ROW - 3, summary_row + 1):
        for col in range(START_COLUMN, START_COLUMN + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thick_border_left
        for col in range(END_COLUMN, END_COLUMN + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thick_border_right

    thin_border = Border(bottom=Side(style="thin", color="000000"))
    thick_border = Border(bottom=Side(style="thick", color="000000"))

    for col in range(START_COLUMN, END_COLUMN + 1):
        ws.cell(row=START_DATA_ROW - 4, column=col).border = ws.cell(row=START_DATA_ROW - 4, column=col).border + thin_border
        ws.cell(row=START_DATA_ROW - 1, column=col).border = ws.cell(row=START_DATA_ROW - 1, column=col).border + thin_border
        ws.cell(row=START_DATA_ROW - 2, column=col).border = ws.cell(row=START_DATA_ROW - 2, column=col).border + thick_border
        ws.cell(row=summary_row - 1, column=col).border = ws.cell(row=summary_row - 1, column=col).border + thin_border
        ws.cell(row=summary_row, column=col).border = ws.cell(row=summary_row, column=col).border + thick_border

def main():
    args = parse_args()
    folder_path = os.getcwd()
    excel_files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx')]

    if len(excel_files) == 0:
        logging.error("No Excel files found in the current directory.")
        return

    for file_name in excel_files:
        file_path = os.path.join(folder_path, file_name)
        process_excel_file(file_path, args.manual)


if __name__ == "__main__":
    main()
